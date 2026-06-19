"""
Assignment 2 - WhatsApp Message Sender + Smart Data Extractor

Use only with contacts who have agreed to receive messages.
WhatsApp Web changes its HTML occasionally. If a selector stops working,
use Playwright Inspector (`PWDEBUG=1`) to update the selector candidates.
"""

from __future__ import annotations

import json
import random
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import (
    BrowserContext,
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


# -----------------------------------------------------------------------------
# Project settings
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
PROFILE_DIR = BASE_DIR / "whatsapp_profile"
TODAY = datetime.now().strftime("%Y-%m-%d")
SCREENSHOT_DIR = BASE_DIR / "screenshots" / TODAY
JSON_REPORT = BASE_DIR / f"whatsapp_report_{TODAY}.json"
EXCEL_REPORT = BASE_DIR / f"whatsapp_report_{TODAY}.xlsx"

DEFAULT_MESSAGE = "Hello {name}, this is a test message from my Playwright assignment."
LOGIN_TIMEOUT_MS = 60_000
ELEMENT_TIMEOUT_MS = 20_000
SEND_CONFIRM_TIMEOUT_MS = 20_000
SCRIPT_VERSION = "1-minute-ready-check-v1"

# Keep this False for the final assignment demo.
# Set it to True while testing selectors; the script will prepare messages but
# will not press Enter.
DRY_RUN = False


# Selector candidates are intentionally grouped in one place because WhatsApp
# Web may change its internal attributes over time.
SEARCH_BOX_SELECTORS = [
    # Newer WhatsApp Web layouts usually keep the chat-search textbox in #side.
    '#side div[contenteditable="true"][role="textbox"]',
    '#side [role="textbox"][contenteditable="true"]',
    '#side input[type="text"]',
    'div[contenteditable="true"][aria-label="Search input textbox"]',
    'div[contenteditable="true"][aria-placeholder*="Search"]',
    'div[contenteditable="true"][role="textbox"][aria-label*="Search"]',
    'div[role="textbox"][aria-label*="Search"]',
    'div[contenteditable="true"][data-tab="3"]',
]

# Do not use only the search box to decide that login is complete. WhatsApp can
# change the search field attributes while the chat list itself is already ready.
WHATSAPP_READY_SELECTORS = [
    '#pane-side',
    '#side',
    '[aria-label="Chat list"]',
    '[data-testid="chat-list"]',
]

MESSAGE_BOX_SELECTORS = [
    '#main footer div[contenteditable="true"][role="textbox"]',
    '#main footer div[contenteditable="true"][data-tab]',
    'footer div[contenteditable="true"][role="textbox"]',
    '#main footer div[contenteditable="true"]',
]

SEND_BUTTON_SELECTORS = [
    '#main footer button[aria-label="Send"]',
    '#main footer button:has(span[data-icon="send"])',
    '#main footer span[data-icon="send"]',
    '#main footer [data-testid="send"]',
]

OUTGOING_MESSAGE_SELECTORS = [
    '#main div.message-out',
    '#main [class*="message-out"]',
    '#main [data-id^="true_"]',
]

INCOMING_MESSAGE_SELECTORS = [
    '#main div.message-in',
    '#main [class*="message-in"]',
    '#main [data-id^="false_"]',
]

MESSAGE_TEXT_SELECTORS = [
    '#main span.selectable-text',
    '#main [data-testid="msg-text"]',
]

CHAT_ROW_SELECTORS = [
    '#pane-side [role="listitem"]',
    '#pane-side [role="row"]',
]


# -----------------------------------------------------------------------------
# Excel helpers
# -----------------------------------------------------------------------------
def create_contacts_template() -> None:
    """Create contacts.xlsx when it does not already exist."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"

    sheet.append(["Name", "Phone", "Message"])
    sheet.append(
        [
            "Test Contact",
            "+15551234567",
            "Hello {name}, this is a test message from my Playwright assignment.",
        ]
    )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 70
    sheet.freeze_panes = "A2"

    # Phone numbers must be saved as text so Excel does not remove + or alter
    # long numbers.
    for row in range(2, 102):
        sheet.cell(row=row, column=2).number_format = "@"
        sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

    workbook.save(CONTACTS_FILE)


def clean_phone(value: Any) -> str:
    """Convert an Excel phone value into a clean string."""
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    phone = str(value).strip()
    if phone.endswith(".0") and phone[:-2].isdigit():
        phone = phone[:-2]
    return phone


def read_contacts() -> list[dict[str, str]]:
    """Read Name, Phone, and Message columns from contacts.xlsx."""
    workbook = load_workbook(CONTACTS_FILE, data_only=True)
    sheet = workbook.active

    headers: dict[str, int] = {}
    for column_number, cell in enumerate(sheet[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = column_number

    required = {"name", "phone", "message"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(
            "contacts.xlsx is missing these columns: " + ", ".join(sorted(missing))
        )

    contacts: list[dict[str, str]] = []
    for row_number in range(2, sheet.max_row + 1):
        name_value = sheet.cell(row_number, headers["name"]).value
        phone_value = sheet.cell(row_number, headers["phone"]).value
        message_value = sheet.cell(row_number, headers["message"]).value

        name = "" if name_value is None else str(name_value).strip()
        phone = clean_phone(phone_value)
        message = "" if message_value is None else str(message_value).strip()

        # Ignore completely empty rows.
        if not name and not phone and not message:
            continue

        if not name and not phone:
            print(f"Skipping Excel row {row_number}: no Name or Phone.")
            continue

        contacts.append(
            {
                "name": name,
                "phone": phone,
                "message_template": message or DEFAULT_MESSAGE,
                "excel_row": str(row_number),
            }
        )

    return contacts


# -----------------------------------------------------------------------------
# Playwright helpers
# -----------------------------------------------------------------------------
def random_delay(page: Page, minimum_ms: int = 2_000, maximum_ms: int = 5_000) -> None:
    """Wait for a random 2-5 seconds using Playwright's wait_for_timeout."""
    delay = random.randint(minimum_ms, maximum_ms)
    print(f"  Human-like wait: {delay / 1000:.1f} seconds")
    page.wait_for_timeout(delay)


def first_visible_locator(
    page: Page,
    selectors: list[str],
    timeout_ms: int = ELEMENT_TIMEOUT_MS,
) -> Locator:
    """Wait for any candidate selector and return its first visible element."""
    combined_selector = ", ".join(selectors)
    page.wait_for_selector(combined_selector, state="visible", timeout=timeout_ms)

    for selector in selectors:
        locator = page.locator(selector)
        for index in range(min(locator.count(), 10)):
            candidate = locator.nth(index)
            if candidate.is_visible():
                return candidate

    raise RuntimeError(f"No visible element found for selectors: {selectors}")


def any_visible(page: Page, selectors: list[str]) -> bool:
    """Return True when at least one selector has a visible element."""
    for selector in selectors:
        try:
            locator = page.locator(selector)
            for index in range(min(locator.count(), 5)):
                if locator.nth(index).is_visible():
                    return True
        except Exception:
            continue
    return False


def wait_for_whatsapp_login(page: Page) -> None:
    """Proceed as soon as WhatsApp is usable; never wait more than 60 seconds."""
    print("Opening WhatsApp Web...")
    page.goto(
        "https://web.whatsapp.com",
        wait_until="domcontentloaded",
        timeout=LOGIN_TIMEOUT_MS,
    )

    print("If a QR code appears, scan it with your phone.")
    print("Checking every 0.5 seconds; maximum wait is 60 seconds...")

    elapsed_ms = 0
    last_progress_second = -10
    qr_message_printed = False

    while elapsed_ms < LOGIN_TIMEOUT_MS:
        # An already-open chat is immediately usable.
        message_box_ready = any_visible(page, MESSAGE_BOX_SELECTORS)

        # The normal logged-in landing page is usable when the left panel exists
        # and either its search field or at least one chat row is visible.
        side_ready = any_visible(page, WHATSAPP_READY_SELECTORS)
        search_ready = any_visible(page, SEARCH_BOX_SELECTORS)
        chat_row_ready = any_visible(page, CHAT_ROW_SELECTORS)

        if message_box_ready or (side_ready and (search_ready or chat_row_ready)):
            page.wait_for_timeout(1_000)
            print(f"WhatsApp Web is ready after {elapsed_ms / 1000:.1f} seconds.")
            return

        # Tell the user when the QR/login page is what Playwright can see.
        qr_selectors = [
            'canvas[aria-label*="Scan"]',
            'div[data-ref]',
            '[data-testid="qrcode"]',
            'text=/link with phone number/i',
        ]
        if not qr_message_printed and any_visible(page, qr_selectors):
            print("QR/login screen detected. Complete the login on your phone.")
            qr_message_printed = True

        current_second = elapsed_ms // 1000
        if current_second - last_progress_second >= 10 and current_second > 0:
            remaining = max(0, (LOGIN_TIMEOUT_MS - elapsed_ms) // 1000)
            print(f"Still waiting for a usable WhatsApp screen... {remaining}s remaining")
            last_progress_second = current_second

        page.wait_for_timeout(500)
        elapsed_ms += 500

    debug_path = BASE_DIR / "whatsapp_login_debug.png"
    page.screenshot(path=str(debug_path), full_page=False)
    print(f"Current browser URL: {page.url}")
    print(f"Saved diagnostic screenshot: {debug_path.name}")
    raise RuntimeError(
        "WhatsApp was not usable within 60 seconds. "
        "The script checked the chat list, search box, and open-chat message box."
    )


def clear_and_type(locator: Locator, text: str, page: Page) -> None:
    """Clear a textbox/contenteditable and type with a human-like key delay."""
    locator.click()
    locator.press("Control+A")
    locator.press("Backspace")
    page.keyboard.type(text, delay=random.randint(60, 120))


def locate_search_result(page: Page, name: str, phone: str) -> Locator:
    """Find a matching chat row after entering a search term."""
    combined_rows = ", ".join(CHAT_ROW_SELECTORS)

    try:
        page.wait_for_selector(combined_rows, state="visible", timeout=12_000)
    except PlaywrightTimeoutError as exc:
        no_result = page.get_by_text(re.compile(r"no .*found|no results", re.I))
        if no_result.count() and no_result.first.is_visible():
            raise RuntimeError("Contact not found in WhatsApp search.") from exc
        raise RuntimeError("No WhatsApp search result appeared.") from exc

    rows = page.locator(combined_rows)

    # Prefer a row containing the contact name.
    if name:
        name_match = rows.filter(has_text=name)
        for index in range(min(name_match.count(), 5)):
            candidate = name_match.nth(index)
            if candidate.is_visible():
                return candidate

    # Then try matching the phone digits. WhatsApp may display spaces or dashes,
    # so compare digits from the row text.
    target_digits = re.sub(r"\D", "", phone)
    for index in range(min(rows.count(), 20)):
        candidate = rows.nth(index)
        if not candidate.is_visible():
            continue
        row_text = candidate.inner_text(timeout=2_000)
        row_digits = re.sub(r"\D", "", row_text)
        if target_digits and target_digits[-7:] in row_digits:
            return candidate

    # Search should already have filtered the list. As the last safe fallback,
    # use the first visible result only when there is no explicit no-results text.
    no_result = page.get_by_text(re.compile(r"no .*found|no results", re.I))
    if no_result.count() and no_result.first.is_visible():
        raise RuntimeError("Contact not found in WhatsApp search.")

    for index in range(min(rows.count(), 10)):
        candidate = rows.nth(index)
        if candidate.is_visible():
            return candidate

    raise RuntimeError("Contact not found in WhatsApp search.")


def wait_for_chat_message_box(page: Page, timeout_ms: int = ELEMENT_TIMEOUT_MS) -> Locator:
    """Confirm that a chat is open by finding its message editor."""
    return first_visible_locator(page, MESSAGE_BOX_SELECTORS, timeout_ms=timeout_ms)


def open_contact_by_phone_url(page: Page, phone: str) -> str:
    """Open an unsaved WhatsApp number directly as a safe fallback."""
    phone = phone.strip()
    if not phone.startswith("+"):
        raise RuntimeError(
            "Phone number must include + and the country code, for example +918553595200."
        )

    digits = re.sub(r"\D", "", phone)
    if len(digits) < 8:
        raise RuntimeError(
            "Phone number is incomplete. Use the full country code, for example +918553595200."
        )

    print("  Search result was not usable; opening the phone number directly.")
    page.goto(
        f"https://web.whatsapp.com/send?phone={quote(digits)}",
        wait_until="domcontentloaded",
    )

    elapsed = 0
    while elapsed < 30_000:
        invalid = page.get_by_text(
            re.compile(
                r"phone number shared via url is invalid|not on whatsapp|couldn't find|isn't on whatsapp",
                re.I,
            )
        )
        try:
            if invalid.count() and invalid.first.is_visible():
                raise RuntimeError("This phone number is not registered on WhatsApp.")
        except RuntimeError:
            raise
        except Exception:
            pass

        try:
            wait_for_chat_message_box(page, timeout_ms=1_000)
            random_delay(page)
            return phone
        except Exception:
            page.wait_for_timeout(500)
            elapsed += 1_500

    raise RuntimeError("WhatsApp could not open this phone number's chat.")


def open_contact(page: Page, name: str, phone: str) -> str:
    """Search by phone/name, open the chat, and fall back to a direct phone URL."""
    search_value = phone or name
    if not search_value:
        raise ValueError("Both Name and Phone are empty.")

    try:
        search_box = first_visible_locator(page, SEARCH_BOX_SELECTORS)
        clear_and_type(search_box, search_value, page)
        random_delay(page)

        result = locate_search_result(page, name=name, phone=phone)
        result.click()

        # The message editor is a more reliable proof of an open chat than
        # WhatsApp's frequently changing #main header markup.
        wait_for_chat_message_box(page)
        random_delay(page)
        return search_value
    except Exception as search_error:
        if phone:
            try:
                return open_contact_by_phone_url(page, phone)
            except Exception as direct_error:
                raise RuntimeError(
                    f"Could not open contact by search ({search_error}) or phone URL ({direct_error})."
                ) from direct_error
        raise


def get_message_box_text(message_box: Locator) -> str:
    """Return text from either a contenteditable element or an input."""
    try:
        tag_name = message_box.evaluate("el => el.tagName.toLowerCase()")
        if tag_name in {"input", "textarea"}:
            return message_box.input_value().strip()
    except Exception:
        pass

    try:
        return message_box.inner_text().strip()
    except Exception:
        return ""


def get_outgoing_count(page: Page) -> int:
    return page.locator(", ".join(OUTGOING_MESSAGE_SELECTORS)).count()


def find_message_text(page: Page, message: str) -> Locator | None:
    """Find a visible copy of the exact/near-exact message in the open chat."""
    all_text = page.locator(", ".join(MESSAGE_TEXT_SELECTORS))
    normalized_target = re.sub(r"\s+", " ", message).strip()

    for index in range(all_text.count() - 1, -1, -1):
        candidate = all_text.nth(index)
        try:
            if not candidate.is_visible():
                continue
            text = re.sub(r"\s+", " ", candidate.inner_text()).strip()
            if text == normalized_target or normalized_target in text:
                return candidate
        except Exception:
            continue
    return None


def send_message(page: Page, message: str) -> tuple[Locator | None, str]:
    """Send the exact personalized Message-column value from the current Excel row."""
    message_box = wait_for_chat_message_box(page)
    previous_outgoing_count = get_outgoing_count(page)

    message_box.click()
    message_box.press("Control+A")
    message_box.press("Backspace")
    page.keyboard.insert_text(message)
    page.wait_for_timeout(750)

    typed_text = get_message_box_text(message_box)
    if not typed_text:
        raise RuntimeError("The chat opened, but Playwright could not type the Excel message.")

    normalized_typed = re.sub(r"\s+", " ", typed_text).strip()
    normalized_expected = re.sub(r"\s+", " ", message).strip()
    if normalized_expected not in normalized_typed and normalized_typed not in normalized_expected:
        raise RuntimeError(
            "The text typed into WhatsApp did not match the Message column in contacts.xlsx."
        )

    print(f"  Excel message prepared: {normalized_expected[:100]}")
    random_delay(page)

    if DRY_RUN:
        print("  DRY_RUN=True: message prepared but not sent.")
        return None, "Dry run"

    clicked_send = False
    for selector in SEND_BUTTON_SELECTORS:
        locator = page.locator(selector)
        for index in range(min(locator.count(), 5)):
            candidate = locator.nth(index)
            if not candidate.is_visible():
                continue
            try:
                if candidate.evaluate("el => el.tagName.toLowerCase()") == "span":
                    parent_button = candidate.locator("xpath=ancestor::button[1]")
                    if parent_button.count() and parent_button.first.is_visible():
                        parent_button.first.click()
                    else:
                        candidate.click()
                else:
                    candidate.click()
                clicked_send = True
                print("  Clicked the WhatsApp Send button.")
                break
            except Exception:
                continue
        if clicked_send:
            break

    if not clicked_send:
        print("  Send button not found; pressing Enter instead.")
        message_box.press("Enter")

    # WhatsApp changes the outgoing bubble classes often. Confirm using several
    # independent signals: a new outgoing node, the sent text appearing, or the
    # composer becoming empty immediately after the Send action.
    elapsed = 0
    composer_cleared = False
    while elapsed < SEND_CONFIRM_TIMEOUT_MS:
        current_outgoing_count = get_outgoing_count(page)
        sent_text = find_message_text(page, message)
        composer_cleared = get_message_box_text(message_box) == ""

        if current_outgoing_count > previous_outgoing_count:
            sent_bubble = page.locator(", ".join(OUTGOING_MESSAGE_SELECTORS)).last
            print("  Message confirmed by a new outgoing message element.")
            random_delay(page)
            return sent_bubble, "New outgoing element"

        if sent_text is not None and composer_cleared:
            print("  Message confirmed by sent text and cleared message box.")
            random_delay(page)
            return sent_text, "Sent text visible"

        if composer_cleared and elapsed >= 2_000:
            print("  Message accepted: the message box cleared after Send.")
            random_delay(page)
            return None, "Message box cleared"

        page.wait_for_timeout(500)
        elapsed += 500

    raise RuntimeError(
        "Send was attempted, but the message box did not clear and no sent message appeared."
    )


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("_") or "contact"


def save_sent_screenshot(page: Page, sent_bubble: Locator | None, label: str) -> str:
    """Save a screenshot of the sent bubble, falling back to the full chat."""
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%H%M%S")
    screenshot_path = SCREENSHOT_DIR / f"{safe_filename(label)}_{timestamp}.png"

    try:
        if sent_bubble is not None:
            sent_bubble.scroll_into_view_if_needed()
            sent_bubble.screenshot(path=str(screenshot_path))
        else:
            page.locator("#main").screenshot(path=str(screenshot_path))
    except Exception:
        page.screenshot(path=str(screenshot_path), full_page=False)

    return str(screenshot_path.relative_to(BASE_DIR))


def extract_message_text(bubble: Locator) -> str:
    """Extract readable message text from a WhatsApp message bubble."""
    selectable = bubble.locator("span.selectable-text")
    pieces: list[str] = []

    for index in range(selectable.count()):
        text = selectable.nth(index).inner_text().strip()
        if text and text not in pieces:
            pieces.append(text)

    if pieces:
        return " ".join(pieces)

    # Fallback for messages whose text is not in span.selectable-text.
    fallback = bubble.inner_text().strip()
    return re.sub(r"\s+", " ", fallback)


def extract_last_three_incoming(page: Page) -> list[str]:
    """Extract the last three incoming (message-in) messages from the open chat."""
    incoming = page.locator(", ".join(INCOMING_MESSAGE_SELECTORS))
    count = incoming.count()

    messages: list[str] = []
    start = max(0, count - 3)
    for index in range(start, count):
        try:
            text = extract_message_text(incoming.nth(index))
            if text:
                messages.append(text)
        except Exception as exc:
            messages.append(f"[Could not extract message: {exc}]")

    return messages[-3:]


# -----------------------------------------------------------------------------
# Reporting
# -----------------------------------------------------------------------------
def save_json_report(results: list[dict[str, Any]], started_at: str, finished_at: str) -> None:
    payload = {
        "assignment": "WhatsApp Message Sender + Smart Data Extractor",
        "date": TODAY,
        "started_at": started_at,
        "finished_at": finished_at,
        "dry_run": DRY_RUN,
        "total_contacts": len(results),
        "sent_count": sum(1 for item in results if item["status"] == "Sent"),
        "failed_count": sum(1 for item in results if item["status"] == "Failed"),
        "results": results,
    }

    with JSON_REPORT.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2, ensure_ascii=False)


def save_excel_report(results: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"

    headers = [
        "Name",
        "Phone",
        "Search Value",
        "Personalized Message",
        "Status",
        "Send Confirmation",
        "Error",
        "Screenshot",
        "Last 3 Incoming Messages",
        "Processed At",
    ]
    sheet.append(headers)

    for result in results:
        last_messages = "\n---\n".join(result.get("last_3_incoming_messages", []))
        sheet.append(
            [
                result.get("name", ""),
                result.get("phone", ""),
                result.get("search_value", ""),
                result.get("personalized_message", ""),
                result.get("status", ""),
                result.get("send_confirmation", ""),
                result.get("error", ""),
                result.get("screenshot", ""),
                last_messages,
                result.get("processed_at", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 22,
        "B": 20,
        "C": 22,
        "D": 55,
        "E": 12,
        "F": 24,
        "G": 38,
        "H": 42,
        "I": 65,
        "J": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Preserve phone formatting as text.
        row[1].number_format = "@"

        status = row[4].value
        if status == "Sent":
            row[4].fill = PatternFill("solid", fgColor="C6EFCE")
        elif status == "Failed":
            row[4].fill = PatternFill("solid", fgColor="FFC7CE")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(EXCEL_REPORT)


def save_reports(results: list[dict[str, Any]], started_at: str) -> None:
    finished_at = datetime.now().isoformat(timespec="seconds")
    save_json_report(results, started_at, finished_at)
    save_excel_report(results)
    print(f"\nJSON report:  {JSON_REPORT.name}")
    print(f"Excel report: {EXCEL_REPORT.name}")


# -----------------------------------------------------------------------------
# Main program
# -----------------------------------------------------------------------------
def process_contact(page: Page, contact: dict[str, str], position: int, total: int) -> dict[str, Any]:
    name = contact["name"]
    phone = contact["phone"]
    template = contact["message_template"]
    personalized_message = template.replace("{name}", name or "there")
    label = name or phone

    result: dict[str, Any] = {
        "name": name,
        "phone": phone,
        "search_value": phone or name,
        "message_template": template,
        "personalized_message": personalized_message,
        "status": "Failed",
        "error": "",
        "screenshot": "",
        "last_3_incoming_messages": [],
        "send_confirmation": "",
        "processed_at": datetime.now().isoformat(timespec="seconds"),
    }

    print(f"\n[{position}/{total}] Processing: {label}")

    try:
        search_value = open_contact(page, name=name, phone=phone)
        result["search_value"] = search_value

        sent_bubble, confirmation = send_message(page, personalized_message)
        result["send_confirmation"] = confirmation
        result["screenshot"] = save_sent_screenshot(page, sent_bubble, label)
        result["last_3_incoming_messages"] = extract_last_three_incoming(page)
        result["status"] = "Dry Run" if DRY_RUN else "Sent"
        print(f"  Status: {result['status']}")

    except Exception as exc:
        result["status"] = "Failed"
        result["error"] = str(exc)
        print(f"  Status: Failed - {exc}")

        # Capture the screen for troubleshooting without crashing the whole run.
        try:
            SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
            error_path = SCREENSHOT_DIR / (
                f"ERROR_{safe_filename(label)}_{datetime.now().strftime('%H%M%S')}.png"
            )
            page.screenshot(path=str(error_path), full_page=False)
            result["screenshot"] = str(error_path.relative_to(BASE_DIR))
        except Exception:
            pass

    return result


def main() -> int:
    print("=" * 68)
    print("WhatsApp Message Sender + Smart Data Extractor")
    print(f"Script version: {SCRIPT_VERSION}")
    print("=" * 68)

    if not CONTACTS_FILE.exists():
        create_contacts_template()
        print(f"Created Excel template: {CONTACTS_FILE.name}")
        print("Open contacts.xlsx, replace the sample row with consented contacts,")
        print("save the file, and run this program again.")
        return 0

    try:
        contacts = read_contacts()
    except Exception as exc:
        print(f"Could not read contacts.xlsx: {exc}")
        return 1

    if not contacts:
        print("No usable contacts were found in contacts.xlsx.")
        return 1

    print(f"Loaded {len(contacts)} contact(s).")
    started_at = datetime.now().isoformat(timespec="seconds")
    results: list[dict[str, Any]] = []

    with sync_playwright() as playwright:
        context: BrowserContext | None = None
        try:
            # Persistent context keeps the WhatsApp login between runs.
            context = playwright.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=False,
                viewport=None,
                args=["--start-maximized"],
            )

            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(ELEMENT_TIMEOUT_MS)
            wait_for_whatsapp_login(page)

            for position, contact in enumerate(contacts, start=1):
                result = process_contact(page, contact, position, len(contacts))
                results.append(result)

                # Save progress after each contact so a later failure does not
                # lose earlier results.
                save_reports(results, started_at)
                random_delay(page)

        except PlaywrightTimeoutError as exc:
            print(f"Playwright timed out: {exc}")
        except KeyboardInterrupt:
            print("\nRun stopped by the user. Saving completed results...")
        except Exception as exc:
            print(f"Unexpected error: {exc}")
        finally:
            save_reports(results, started_at)
            if context is not None:
                context.close()

    print("\nRun complete.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
